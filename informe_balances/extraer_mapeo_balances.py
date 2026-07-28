# -*- coding: utf-8 -*-
"""
Extrae el "mapeo maestro" de los dos balances (pesos y dolares) a un JSON
parametrizado. Es la base de conocimiento del Informe C (Balance Pesos y
Dolares): por cada archivo, donde esta cada cosa y que celda alimenta a que
linea de los estados contables.
"""
import json, re, sys
import openpyxl
from openpyxl.utils import get_column_letter

def norm(s):
    if s is None: return ''
    return re.sub(r'\s+', ' ', str(s)).strip()

CODE_RE = re.compile(r'^\s*(\d{5,})\s*-\s*(.+?)\s*$')
SALDOS_REF = None  # se setea por workbook

def extraer(config):
    wb = openpyxl.load_workbook(config['archivo'], data_only=False)
    out = {
        'moneda': config['moneda'],
        'params': {
            'hoja1_col_clave': config['hoja1_col_clave'],
            'hoja1_col_valor': config['hoja1_col_valor'],
            'saldos_cols_cuenta': config['saldos_cols_cuenta'],
            'saldos_col_valor': config['saldos_col_valor'],
        },
    }

    # ---- 1. SALDOS: cuentas, filas, capitulos ----
    ws = wb['SALDOS']
    cuentas = {}
    duplicadas = []
    capitulos = {}
    cap_actual = None
    for r in range(1, ws.max_row + 1):
        # capitulo: puede estar en cualquiera de las columnas de texto
        for c_cap in config['saldos_cols_capitulo']:
            v = norm(ws[f'{c_cap}{r}'].value)
            if v in ('ACTIVO', 'PASIVO', 'PATRIMONIO NETO', 'RESULTADOS'):
                cap_actual = v
                capitulos[v] = {'desde_fila': r}
        # cuenta: buscar el patron "codigo - nombre" en las columnas candidatas
        for c_cta in config['saldos_cols_cuenta']:
            v = norm(ws[f'{c_cta}{r}'].value)
            m = CODE_RE.match(v)
            if m:
                cod = m.group(1)
                if cod in cuentas and cuentas[cod]['fila_saldos'] != r:
                    duplicadas.append({'codigo': cod, 'filas': [cuentas[cod]['fila_saldos'], r]})
                else:
                    cuentas[cod] = {
                        'fila_saldos': r,
                        'label': m.group(2),
                        'capitulo': cap_actual,
                        'col_texto': c_cta,
                    }
                break  # con una columna que matchee alcanza para esta fila
    out['cuentas_saldos'] = cuentas
    out['cuentas_duplicadas'] = duplicadas
    out['capitulos'] = capitulos

    # ---- 2. referencias: que linea de cada hoja de estados usa cada fila de SALDOS ----
    saldo_col = config['saldos_col_valor']
    ref_re = re.compile(r"SALDOS!\$?" + saldo_col + r"\$?(\d+)")
    fila_a_codigo = {info['fila_saldos']: cod for cod, info in cuentas.items()}
    referencias = []
    for hoja in ['Activo y Pasivo', 'Anexo II', 'Anexo I', 'Resultados', 'Balance', 'Pat.Neto']:
        if hoja not in wb.sheetnames: continue
        wsx = wb[hoja]
        for r in range(1, wsx.max_row + 1):
            for c in range(1, wsx.max_column + 1):
                v = wsx.cell(row=r, column=c).value
                if isinstance(v, str) and v.startswith('=') and 'SALDOS!' in v:
                    filas = [int(x) for x in ref_re.findall(v)]
                    if not filas: continue
                    # etiqueta de la linea: buscar texto hacia la izquierda en la misma fila
                    etiqueta = None
                    for cc in range(c - 1, 0, -1):
                        t = wsx.cell(row=r, column=cc).value
                        if isinstance(t, str) and norm(t) not in ('', '-'):
                            etiqueta = norm(t); break
                    referencias.append({
                        'hoja': hoja,
                        'celda': f'{get_column_letter(c)}{r}',
                        'etiqueta_linea': etiqueta,
                        'filas_saldos': filas,
                        'cuentas': [fila_a_codigo.get(f) for f in filas],
                        'formula': v,
                    })
    out['referencias_estados'] = referencias

    # ---- 3. lineas clasificables (para el formulario de cuenta nueva) ----
    # Anexo II: conceptos (col B) con sus 4 columnas F/G/H/I
    wsx = wb['Anexo II']
    conceptos = []
    for r in range(10, 102):
        b = norm(wsx[f'B{r}'].value)
        if b and b not in ('', '-'):
            conceptos.append({'fila': r, 'concepto': b})
    out['anexo2_conceptos'] = conceptos
    out['anexo2_columnas'] = {'F': 'ADMINISTRACION', 'G': 'COMERCIALIZACION',
                              'H': 'EXPLORACION', 'I': 'FINANCIEROS'}

    # Activo y Pasivo: rubros (filas con "a." "b." etc en col B o subtotales SUM)
    wsx = wb['Activo y Pasivo']
    rubros = []
    for r in range(1, wsx.max_row + 1):
        bb = norm(wsx[f'B{r}'].value)
        cc = norm(wsx[f'C{r}'].value)
        if re.match(r'^[a-z]\.$', bb) and cc:
            rubros.append({'fila_titulo': r, 'rubro': cc})
    out['nota4_rubros'] = rubros

    return out


if __name__ == '__main__':
    configs = [
        {
            'moneda': 'dolares',
            'archivo': 'SCA_Balance_al_06-2026_dolares_.xlsx',
            'hoja1_col_clave': 'A', 'hoja1_col_valor': 'D',
            'saldos_cols_cuenta': ['B', 'A'], 'saldos_col_valor': 'C',
            'saldos_cols_capitulo': ['A'],
        },
        {
            'moneda': 'pesos',
            'archivo': 'SCA_Balance_a_06-2026_Pesos.xlsx',
            'hoja1_col_clave': 'A', 'hoja1_col_valor': 'E',
            'saldos_cols_cuenta': ['C', 'D', 'E'], 'saldos_col_valor': 'G',
            'saldos_cols_capitulo': ['C', 'D', 'A'],
        },
    ]
    resultado = {}
    for cfg in configs:
        m = extraer(cfg)
        resultado[cfg['moneda']] = m
        print(f"{cfg['moneda']}: {len(m['cuentas_saldos'])} cuentas en SALDOS, "
              f"{len(m['referencias_estados'])} referencias en estados, "
              f"{len(m['anexo2_conceptos'])} conceptos Anexo II, "
              f"{len(m['nota4_rubros'])} rubros Nota 4")
    with open('mapeo_balances_pesos_dolares.json', 'w', encoding='utf-8') as f:
        json.dump(resultado, f, ensure_ascii=False, indent=2)
    print('Guardado mapeo_balances_pesos_dolares.json')
